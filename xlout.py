from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill, Color
from openpyxl.utils import get_column_letter
#from openpyxl.worksheet.cell_range import max_row

def srs2xl(reqIdDic,reqIdDic_usecase):
    wb=Workbook()
    ws1 = wb.active
    ws1.title = "Input_Sheet"
    insert_cell(1,1,"Word",ws1,border=True)
    insert_cell(1,2,"ReqID",ws1,border=True)
    color_cell(1,1,ws1)
    color_cell(1,2,ws1)
    reqIdDic_merged = reqIdDic+reqIdDic_usecase
    for i in range(2,len(reqIdDic_merged)) :
        for j in range(1,3) :
            c = ws1.cell(i,j)
            if reqIdDic_merged[i-2][1] != None :
                c.value = reqIdDic_merged[i-2][j-1]
    wb.save('SRS 분석 결과.xlsx')

# def srs2xl(reqIdDic,reqIdDic_usecase):
#     dic_xlsx = load_workbook('Dictionary.xlsx')
#     ws1 = dic_xlsx.create_sheet('식별자 Dictionary')
#     insert_cell(1,1,"Word",ws1,border=True)
#     insert_cell(1,2,"ReqID",ws1,border=True)
#     color_cell(1,1,ws1)
#     color_cell(1,2,ws1)
#     reqIdDic_merged = reqIdDic+reqIdDic_usecase
#     for i in range(2,len(reqIdDic_merged)) :
#         for j in range(1,3) :
#             c = ws1.cell(i,j)
#             if reqIdDic_merged[i-2][1] != None :
#                 c.value = reqIdDic_merged[i-2][j-1]
#     dic_xlsx.save('Dictionary.xlsx')

def insert_cell(r,c,v,ws1,border=False):
    cell = ws1.cell(r,c)
    cell.value = v
    if border == True:
        cell.font = Font(size=10)
        cell.border = Border(left=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
        cell.alignment=Alignment(horizontal='center',vertical='center')

def find_cusor(Dp_R,Dp_C,ws1):
    cell = ws1.cell(Dp_R,Dp_C)
    while(True):
        while(True):
            if cell.value!=None :
                Dp_C+=1
                cell = ws1.cell(Dp_R,Dp_C)
            else :
                Dp_C = 1
                Dp_R +=1
                cell = ws1.cell(Dp_R,Dp_C)
                break
        check_the_cell = ws1.cell(Dp_R+1,Dp_C+1)
        if cell.value == None and check_the_cell.value == None :
            return Dp_R,Dp_C

class indent_number:
    def __init__(self,affIndent=''):
        self.aff = affIndent
        self.c_ind = 0
    def merge(self):
        print(self.aff + '.'+ str(self.c_ind))
        return self.aff + '.'+ str(self.c_ind)

class component_word:
    def __init__(self,me,template,child=None):
        self.me = me
        self.template =template
        self.children = []
        self.ilvl = 0
        if child != None:
            self.insert_child(child)
    def insert_child(self,child):
        self.children.append(child)

class is_a_word:
    def __init__(self,me,template,child=None):
        self.me = me
        self.template =template
        self.children = []
        self.ilvl = 0
        if child != None:
            self.insert_child(child)
    def insert_child(self,child):
        self.children.append(child)

def setup_TitleCell(row,col,ws1,wb):
    sheet=wb.get_sheet_by_name('testcase_input')
    sheet.merge_cells('A1:E1')
    #ws1.merge_cell(A1:E1)
    cell = ws1.cell(row,col)
    cell.value = 'Testcase_Input'
    cell.font=Font(size=15,bold=True)
    cell.border = Border(left=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    cell.alignment=Alignment(horizontal='center',vertical='center')
    color_cell(1,1,ws1)

def list_merge(list):
    out = str()
    for i in list:
        out += i
    return out

def color_cell(row,col,ws1):
    cell = ws1.cell(row,col)
    cell.fill = PatternFill(patternType='solid',fgColor=Color('C0C0C0'))

def srs_out(final_srs,doc_name):
    max_indent = 10
    reqId_col=1
    wb=Workbook()
    ws1 = wb.active
    ws1.title = 'testcase_input'
    row = 2
    col = 1
    setup_TitleCell(1,1,ws1,wb)
    insert_cell(row,reqId_col,'문서명',ws1,border=True)
    color_cell(row,reqId_col,ws1)
    insert_cell(row,reqId_col+1,doc_name,ws1,border=True)
    row+=1
    for reqId in list(final_srs.keys()):
        row += 1
        insert_cell(row,reqId_col,'식별자',ws1,border=True)
        color_cell(row,reqId_col,ws1)
        insert_cell(row,reqId_col+1,reqId,ws1,border=True)
        row += 1
        insert_cell(row,reqId_col,'요구사항',ws1,border=True)
        color_cell(row,reqId_col,ws1)
        row += 1
        p_ilvl = 1
        initflag = 1
        ind_list=[indent_number() for x in range(max_indent)]
        component_word_list=[]
        is_a_word_list=[]
        p_template = None
        is_a_cnt = 0
        has_a_cnt = 0
        current_stack = []
        o_ds_list = final_srs.get(reqId)
        for o_ds in o_ds_list:
            p_col = reqId_col
            c_ilvl=o_ds.ilvl
            c_template = o_ds.template
            if c_template == 'has_a' :
                try:
                    has_a_cnt += 1
                    component_word_list.append(component_word(o_ds.words[-1],'has_a'))
                except:
                    print('component word range error')
                    pass
            elif c_template == 'is_a' :
                try:
                    is_a_cnt += 1
                    is_a_word_list.append(is_a_word(o_ds.words[-1],'is_a'))
                except:
                    print('is a word range error')
                    pass
            if initflag == 1:
                ind_list[c_ilvl].c_ind = 1
                indent = '1'
                insert_cell(row,p_col,indent,ws1)
                p_col += 1
                p_ilvl = c_ilvl
                initflag=0
                for w in o_ds.words:
                    insert_cell(row,p_col,w,ws1)
                    p_col += 1
                row +=1
            elif c_ilvl > p_ilvl :
                ind_list[c_ilvl].c_ind = 1
                if ind_list[p_ilvl].aff == '':
                    ind_list[c_ilvl].aff = str(ind_list[p_ilvl].c_ind) + '.'
                elif ind_list[p_ilvl].aff != '':
                    ind_list[c_ilvl].aff = ind_list[p_ilvl].aff + str(ind_list[p_ilvl].c_ind) + '.'
                indent = ind_list[c_ilvl].aff + str(ind_list[c_ilvl].c_ind)
                #insert_cell(row,p_col,indent,ws1)
                p_ilvl = c_ilvl
                #p_col += 1
                if (c_template == 'not_sentence' and p_template == 'has_a'):
                    component_word_list[-1].ilvl = c_ilvl
                    component_word_list[-1].insert_child(list_merge(o_ds.words))
                    current_stack.append(component_word_list[-1])
                elif (c_template == 'not_sentence' and p_template == 'is_a'):
                    is_a_word_list[-1].ilvl = c_ilvl
                    is_a_word_list[-1].insert_child(list_merge(o_ds.words))
                    current_stack.append(is_a_word_list[-1])
                elif(c_template == 'not_sentence' and p_template == 'not_sentence'):
                    c_word = current_stack[-1].children[-1]
                    is_a_cnt += 1
                    is_a_word_list.append(is_a_word(c_word,'is_a'))
                    is_a_word_list[-1].ilvl = c_ilvl
                    is_a_word_list[-1].insert_child(list_merge(o_ds.words))
                    current_stack.append(is_a_word_list[-1])
                else :
                    insert_cell(row,p_col,indent,ws1)
                    p_col += 1
                    for w in o_ds.words:
                        insert_cell(row,p_col,w,ws1)
                        p_col += 1
                    row +=1
            elif c_ilvl == p_ilvl :
                ind_list[c_ilvl].c_ind += 1
                indent = ind_list[c_ilvl].aff + str(ind_list[c_ilvl].c_ind)
                #insert_cell(row,p_col,indent,ws1)
                p_ilvl = c_ilvl
                #p_col += 1
                if len(current_stack) != 0 :
                    if c_template == 'not_sentence' and current_stack[-1].ilvl == c_ilvl:
                        current_stack[-1].insert_child(list_merge(o_ds.words))
                else :
                    insert_cell(row,p_col,indent,ws1)
                    p_col += 1
                    for w in o_ds.words:
                        insert_cell(row,p_col,w,ws1)
                        p_col += 1
                    row +=1
            elif c_ilvl < p_ilvl :
                ind_list[c_ilvl].c_ind += 1
                indent = ind_list[c_ilvl].aff + str(ind_list[c_ilvl].c_ind)
                #insert_cell(row,p_col,indent,ws1)
                p_ilvl = c_ilvl
                #p_col += 1
                if c_template == 'not_sentence':
                    current_stack.pop()
                    if len(current_stack) != 0 :
                        current_stack[-1].insert_child(list_merge(o_ds.words))
                else :
                    insert_cell(row,p_col,indent,ws1)
                    p_col += 1
                    for w in o_ds.words:
                        insert_cell(row,p_col,w,ws1)
                        p_col += 1
                    row +=1
            # for w in o_ds.words:
            #     insert_cell(row,p_col,w,ws1)
            #     p_col += 1
            # row +=1
            p_template = c_template
        p_col =reqId_col
        insert_cell(row,p_col,'has-a list',ws1,border=True)
        color_cell(row,p_col,ws1)
        p_col = reqId_col
        row += 1
        print('component_word_list')
        print(component_word_list)
        print('component_words')
        for k in component_word_list:
            print(k)
        for c_w in component_word_list:
            insert_cell(row,p_col,c_w.me,ws1)
            p_col +=1
            for w in c_w.children :
                insert_cell(row,p_col,w,ws1)
                p_col +=1
            p_col = reqId_col
            row+=1
        p_col = reqId_col
        insert_cell(row,p_col,'is-a list',ws1,border=True)
        color_cell(row,p_col,ws1)
        p_col = reqId_col
        row += 1
        print('is_a_word_list')
        print(is_a_word_list)
        print('is_a_words')
        for k in is_a_word_list:
            print(k)
        for i_w in is_a_word_list:
            insert_cell(row,p_col,i_w.me,ws1)
            p_col +=1
            for w in i_w.children :
                insert_cell(row,p_col,w,ws1)
                p_col +=1
            p_col = reqId_col
            row+=1
        row+=1
    wb.save('testcase_input_by_srs.xlsx')
    return row

def usecase_out(usecase,doc_name,current_row=1):
    opt = 1
    max_indent = 10
    row = current_row
    offset_row=current_row
    offset_col=1
    wb=load_workbook('testcase_input_by_srs.xlsx')
    ws1 = wb.active
    #row,col=find_cusor(offset_row,offset_col,ws1)
    row += 2
    #insert_cell(row,col,'문서명',ws1,border=True)
    #insert_cell(row,col+1,doc_name,ws1,border=True)
    #row +=1
    offset_row = row
    p_ilvl = 0
    init_flag = 0
    ind_list=[indent_number() for x in range(max_indent)]
    for c in usecase :
        if c == []:
            insert_cell(row,col,'',ws1)
            row += 1
        else:
            ilvl = c[0].get('ilvl')
            if ilvl == 0 :
                if offset_col == 1:
                    offset_row += 1
                    row+=1
                    insert_cell(offset_row,offset_col,'식별자',ws1,border=True)
                    color_cell(offset_row,offset_col,ws1)
                    insert_cell(offset_row,offset_col+1,c[0].get('reqId'),ws1,border=True)
                    offset_row += 1
                    row+=1
                    insert_cell(offset_row,offset_col,"시나리오",ws1,border=True)
                    color_cell(offset_row,offset_col,ws1)
                    offset_row += 1
                    row += 1
                    words = str()
                    for d in c :
                        words += d.get('word') + ' '
                    insert_cell(offset_row,offset_col,words,ws1,border=True)
                    offset_col = 2
                    row += 1
                    col = 1
                    p_ilvl = 0
                    init_flag = 0
                    left_max = 0
                    ind_list=[indent_number() for x in range(max_indent)]
                elif offset_col == 2:
                    words = str()
                    for d in c :
                        words += d.get('word') + ' '
                    insert_cell(offset_row,offset_col+left_max,words,ws1,border=True)
                    col = offset_col + left_max
                    offset_col = 1
                    offset_row,row= row,offset_row
                    row +=1
            else :
                words = str()
                if init_flag == 0:
                    ind_list[ilvl].c_ind = 1
                    indent = '1'
                    p_ilvl = ilvl
                    init_flag = 1
                elif ilvl > p_ilvl:
                    ind_list[ilvl].c_ind = 1
                    if ind_list[p_ilvl].aff == '':
                        ind_list[ilvl].aff = str(ind_list[p_ilvl].c_ind) + '.'
                    elif ind_list[p_ilvl].aff != '':
                        ind_list[ilvl].aff = ind_list[p_ilvl].aff + str(ind_list[p_ilvl].c_ind) + '.'
                    indent = ind_list[ilvl].aff + str(ind_list[ilvl].c_ind)
                    p_ilvl = ilvl
                elif ilvl==p_ilvl:
                    ind_list[ilvl].c_ind += 1
                    indent = ind_list[ilvl].aff + str(ind_list[ilvl].c_ind)
                    p_ilvl = ilvl
                elif ilvl < p_ilvl:
                    ind_list[ilvl].c_ind += 1
                    indent = ind_list[ilvl].aff + str(ind_list[ilvl].c_ind)
                    p_ilvl = ilvl
                words += indent + '. '
                if opt == 0 :
                    for d in c:
                        words += d.get('word') + ' '
                    insert_cell(row,col,words,ws1)
                elif opt == 1 :
                    insert_cell(row,col,words,ws1)
                    col += 1
                    if offset_col == 2:
                        left = 0
                        for d in c:
                            insert_cell(row,col,d.get('word'),ws1)
                            col += 1
                            left += 1
                        if left_max < left:
                            left_max = left
                        col = 1
                    elif offset_col == 1:
                        for d in c:
                            insert_cell(row,col,d.get('word'),ws1)
                            col += 1
                        col = offset_col + left_max + 1
                row += 1
    #change_sizeofcell(ws1,wb)
    wb.save('testcase_input_by_srs.xlsx')